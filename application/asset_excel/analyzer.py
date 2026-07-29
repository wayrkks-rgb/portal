from __future__ import annotations

import os
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import load_workbook

from application.common import col_letter_to_idx
from application.settings import OUTPUT_DIR

def _clean_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y%m%d')
    text = str(value).strip()
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text

def _json_safe(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value

def _parse_col_list(value, default=None):
    if default is None:
        default = []
    if value is None:
        return default
    if isinstance(value, list):
        cols = value
    else:
        cols = str(value).split(',')
    result = []
    for c in cols:
        c = str(c).strip().upper()
        if c:
            result.append(c)
    return result or default

def _normalize_status(value):
    return _clean_value(value).replace(' ', '')

def _eosl_year(value):
    text = _clean_value(value)
    if not text:
        return '확인필요'
    if text.startswith('9999'):
        return '미정'
    year = text[:4]
    if year.isdigit():
        return year
    return '확인필요'

def _cell(row, idx):
    return row[idx] if idx < len(row) else None

def _col_name(headers, idx):
    if 0 <= idx < len(headers):
        name = _clean_value(headers[idx])
        if name:
            return name
    return openpyxl.utils.get_column_letter(idx + 1)

def _load_asset_rows(path, header_row=3, data_start_row=4, key_cols=None):
    """엑셀 자산현황을 key 기준 dict로 변환한다."""
    key_cols = key_cols or ['A']
    key_indexes = [col_letter_to_idx(c) for c in key_cols]

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    max_col = ws.max_column or 1
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, max_col=max_col, values_only=True), ())
    headers = [_clean_value(h) or openpyxl.utils.get_column_letter(i + 1) for i, h in enumerate(list(header_values))]

    records = {}
    duplicate_keys = []
    empty_key_rows = []

    for offset, row in enumerate(ws.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True), start=0):
        row_idx = data_start_row + offset
        values = list(row)
        if not any(v is not None and str(v).strip() != '' for v in values):
            continue

        key_parts = []
        for idx in key_indexes:
            key_parts.append(_clean_value(_cell(values, idx)))
        key = '|'.join(key_parts).strip('|').strip()

        if not key:
            empty_key_rows.append(row_idx)
            continue
        if key in records:
            duplicate_keys.append(key)

        records[key] = {
            'key': key,
            'row_num': row_idx,
            'values': values
        }

    wb.close()
    return {
        'headers': headers,
        'records': records,
        'duplicate_keys': sorted(set(duplicate_keys)),
        'empty_key_rows': empty_key_rows,
        'max_col': max_col,
        'sheet_name': sheet_name
    }

def _asset_brief(record, headers, important_indexes=None):
    if important_indexes is None:
        important_indexes = [0, 1, 5, 26, 34, 35, 36]
    values = record['values']
    brief = {
        'key': record['key'],
        'row_num': record['row_num']
    }
    fields = []
    for idx in important_indexes:
        if idx < len(values):
            fields.append({
                'col': openpyxl.utils.get_column_letter(idx + 1),
                'name': _col_name(headers, idx),
                'value': _clean_value(values[idx]) or '-'
            })
    brief['fields'] = fields
    return brief

def _status_scope_filter(records, status_idx, scope):
    active = {'운영', '대기'}
    inactive = {'미사용', '매각/폐기'}
    if scope == 'active':
        allowed = active
    elif scope == 'inactive':
        allowed = inactive
    else:
        allowed = None

    if allowed is None:
        return list(records.values())

    filtered = []
    for rec in records.values():
        status = _normalize_status(_cell(rec['values'], status_idx))
        if status in {s.replace(' ', '') for s in allowed}:
            filtered.append(rec)
    return filtered

def _count_by(records, col_idx):
    counts = {}
    for rec in records:
        key = _clean_value(_cell(rec['values'], col_idx)) or '확인필요'
        counts[key] = counts.get(key, 0) + 1
    return sorted([{'name': k, 'count': v} for k, v in counts.items()], key=lambda x: (-x['count'], x['name']))

def _summary_for_records(records, os_col, os_ver_col, physical_col, eosl_col):
    os_idx = col_letter_to_idx(os_col)
    os_ver_idx = col_letter_to_idx(os_ver_col)
    physical_idx = col_letter_to_idx(physical_col)
    eosl_idx = col_letter_to_idx(eosl_col)

    eosl_counts = {}
    pivot = {}
    for rec in records:
        values = rec['values']
        os_ver = _clean_value(_cell(values, os_ver_idx)) or '확인필요'
        year = _eosl_year(_cell(values, eosl_idx))
        eosl_counts[year] = eosl_counts.get(year, 0) + 1
        pivot.setdefault(os_ver, {})[year] = pivot.setdefault(os_ver, {}).get(year, 0) + 1

    def sort_year(item):
        y = item['year']
        if y.isdigit():
            return (0, int(y))
        if y == '미정':
            return (1, 9999)
        return (2, 10000)

    eosl_by_year = sorted([{'year': k, 'count': v} for k, v in eosl_counts.items()], key=sort_year)
    eosl_years = [x['year'] for x in eosl_by_year]
    pivot_rows = []
    for os_ver, year_map in pivot.items():
        row = {'os_version': os_ver, 'total': sum(year_map.values()), 'years': {y: year_map.get(y, 0) for y in eosl_years}}
        pivot_rows.append(row)
    pivot_rows.sort(key=lambda x: (-x['total'], x['os_version']))

    return {
        'total_count': len(records),
        'os_counts': _count_by(records, os_idx),
        'os_version_counts': _count_by(records, os_ver_idx),
        'physical_counts': _count_by(records, physical_idx),
        'eosl_by_year': eosl_by_year,
        'eosl_years': eosl_years,
        'eosl_pivot': pivot_rows
    }

def _make_compare_excel(result):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'자산_변경현황_분석_{ts}.xlsx'
    out_path = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '요약'

    summary = result['summary']
    rows = [
        ['구분', '값'],
        ['이전 자산 수', result['compare']['old_count']],
        ['변경 자산 수', result['compare']['new_count']],
        ['추가 자산', result['compare']['added_count']],
        ['상태값 기준 삭제', result['compare']['status_deleted_count']],
        ['일반 변경', result['compare']['changed_count']],
        ['물리적 삭제', result['compare']['removed_count']],
        ['유지 자산', result['compare']['common_count']],
        ['현황 집계 기준', result['options']['scope_label']],
        ['현황 대상 대수', summary['total_count']],
    ]
    for r in rows:
        ws.append(r)

    def write_simple_count_sheet(title, data, key_name='구분'):
        sh = wb.create_sheet(title)
        sh.append([key_name, '대수'])
        for item in data:
            sh.append([item.get('name') or item.get('year'), item['count']])

    def write_asset_sheet(title, items):
        sh = wb.create_sheet(title)
        if not items:
            sh.append(['대상 없음'])
            return
        field_names = []
        seen = set()
        for item in items:
            for f in item.get('fields', []):
                label = f"{f['col']} {f['name']}"
                if label not in seen:
                    field_names.append(label)
                    seen.add(label)
        sh.append(['자산코드', '행번호'] + field_names)
        for item in items:
            fmap = {f"{f['col']} {f['name']}": f.get('value', '') for f in item.get('fields', [])}
            sh.append([item['key'], item['row_num']] + [fmap.get(n, '') for n in field_names])

    def write_change_sheet(title, items):
        sh = wb.create_sheet(title)
        sh.append(['자산코드', '행번호(이전)', '행번호(변경)', '변경열', '항목명', '이전값', '변경값'])
        if not items:
            sh.append(['대상 없음', '', '', '', '', '', ''])
            return
        for item in items:
            for ch in item.get('changes', []):
                sh.append([item['key'], item['old_row_num'], item['new_row_num'], ch['col'], ch['name'], ch['before'], ch['after']])

    write_asset_sheet('추가자산', result['added'])
    write_asset_sheet('상태값기준삭제', result['status_deleted'])
    write_change_sheet('일반변경', result['changed'])
    write_asset_sheet('물리적삭제', result['removed'])
    write_simple_count_sheet('OS별대수', summary['os_counts'], 'OS')
    write_simple_count_sheet('OS버전별대수', summary['os_version_counts'], 'OS VER')
    write_simple_count_sheet('물리논리대수', summary['physical_counts'], '물리/논리')
    write_simple_count_sheet('EOSL연도별대수', summary['eosl_by_year'], 'EOSL 연도')

    sh = wb.create_sheet('OS버전_EOSL')
    years = summary['eosl_years']
    sh.append(['OS VER', '합계'] + years)
    for item in summary['eosl_pivot']:
        sh.append([item['os_version'], item['total']] + [item['years'].get(y, 0) for y in years])

    custom = result.get('custom_summary')
    if custom:
        write_simple_count_sheet('사용자지정집계', custom['counts'], custom['group_label'])

    for sh in wb.worksheets:
        for col in sh.columns:
            max_len = 0
            letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sh.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    wb.save(out_path)
    return filename

