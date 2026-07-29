from __future__ import annotations

import os
from datetime import datetime

import openpyxl
from flask import Blueprint, jsonify, render_template, request, session
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from application.asset_excel.analyzer import (
    _asset_brief, _cell, _clean_value, _col_name, _count_by, _json_safe, _load_asset_rows,
    _make_compare_excel, _normalize_status, _parse_col_list, _status_scope_filter,
    _summary_for_records,
)
from application.common import col_letter_to_idx, load_json, require_login, save_json
from application.settings import HISTORY_FILE, UPLOAD_DIR

bp = Blueprint("asset_excel", __name__)

@bp.route("/compare")
@require_login
def compare():
    return render_template('main.html', user=session['user'], page='compare')

@bp.route("/api/upload-asset", methods=["POST"])
@require_login
def upload_asset():
    """자산현황 파일 업로드 전용 API. 기본 구조: 3행 헤더, 4행 데이터."""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '파일이 없습니다.'}), 400

    header_row = int(request.form.get('header_row', 3))
    data_start_row = int(request.form.get('data_start_row', 4))

    filename = secure_filename(f.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    saved_name = f'asset_{ts}_{filename}'
    path = os.path.join(UPLOAD_DIR, saved_name)
    f.save(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        max_col = ws.max_column or 1
        header = next(ws.iter_rows(min_row=header_row, max_row=header_row, max_col=max_col, values_only=True), ())
        header = list(header)
        preview = []
        row_count = 0
        for values in ws.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True):
            values = list(values)
            if any(v is not None and str(v).strip() != '' for v in values):
                row_count += 1
                if len(preview) < 5:
                    preview.append(values)
        wb.close()

        history = load_json(HISTORY_FILE)
        history.append({
            'id': f'asset_{ts}',
            'type': 'asset',
            'filename': filename,
            'saved_name': saved_name,
            'uploaded_at': datetime.now().isoformat(),
            'uploaded_by': session['user']['name'],
            'row_count': row_count,
            'header_row': header_row,
            'data_start_row': data_start_row
        })
        save_json(HISTORY_FILE, history)

        return jsonify({
            'success': True,
            'file_id': ts,
            'saved_name': saved_name,
            'filename': filename,
            'row_count': row_count,
            'header': [_json_safe(v) for v in header],
            'preview': [[_json_safe(v) for v in r] for r in preview]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route("/api/compare", methods=["POST"])
@require_login
def do_compare():
    data = request.get_json()
    f1 = data.get('file1')
    f2 = data.get('file2')

    if not f1 or not f2:
        return jsonify({'error': '이전 파일과 변경 파일을 모두 선택해야 합니다.'}), 400

    path1 = os.path.join(UPLOAD_DIR, f1)
    path2 = os.path.join(UPLOAD_DIR, f2)
    if not os.path.exists(path1) or not os.path.exists(path2):
        return jsonify({'error': '업로드된 파일을 찾을 수 없습니다.'}), 404

    key_cols = _parse_col_list(data.get('key_cols') or data.get('id_cols'), ['A'])
    header_row = int(data.get('header_row', 3))
    data_start_row = int(data.get('data_start_row', 4))
    status_col = data.get('status_col', 'F').strip().upper()
    status_idx = col_letter_to_idx(status_col)
    exclude_cols = set(_parse_col_list(data.get('exclude_cols'), ['F']))
    scope = data.get('scope', 'active')
    os_col = data.get('os_col', 'AI').strip().upper()
    os_ver_col = data.get('os_version_col', 'AJ').strip().upper()
    physical_col = data.get('physical_col', 'AA').strip().upper()
    eosl_col = data.get('eosl_col', 'AK').strip().upper()

    active_status = {'운영', '대기'}
    inactive_status = {'미사용', '매각/폐기'}
    active_norm = {s.replace(' ', '') for s in active_status}
    inactive_norm = {s.replace(' ', '') for s in inactive_status}
    scope_label_map = {'active': '운영+대기', 'all': '전체', 'inactive': '미사용+매각폐기'}

    try:
        old_data = _load_asset_rows(path1, header_row, data_start_row, key_cols)
        new_data = _load_asset_rows(path2, header_row, data_start_row, key_cols)
        old_records = old_data['records']
        new_records = new_data['records']
        headers = new_data['headers'] if new_data['headers'] else old_data['headers']

        old_keys = set(old_records.keys())
        new_keys = set(new_records.keys())
        added_keys = sorted(new_keys - old_keys)
        removed_keys = sorted(old_keys - new_keys)
        common_keys = sorted(old_keys & new_keys)

        important_indexes = sorted(set([0, 1, status_idx, col_letter_to_idx(physical_col), col_letter_to_idx(os_col), col_letter_to_idx(os_ver_col), col_letter_to_idx(eosl_col)]))
        added = [_asset_brief(new_records[k], headers, important_indexes) for k in added_keys]
        removed = [_asset_brief(old_records[k], old_data['headers'], important_indexes) for k in removed_keys]

        changed = []
        status_deleted = []
        for key in common_keys:
            old_rec = old_records[key]
            new_rec = new_records[key]
            old_values = old_rec['values']
            new_values = new_rec['values']
            old_status = _normalize_status(_cell(old_values, status_idx))
            new_status = _normalize_status(_cell(new_values, status_idx))

            if old_status in active_norm and new_status in inactive_norm:
                brief = _asset_brief(new_rec, headers, important_indexes)
                brief['old_status'] = _clean_value(_cell(old_values, status_idx)) or '-'
                brief['new_status'] = _clean_value(_cell(new_values, status_idx)) or '-'
                brief['old_row_num'] = old_rec['row_num']
                brief['new_row_num'] = new_rec['row_num']
                status_deleted.append(brief)
                continue

            max_len = max(len(old_values), len(new_values), len(headers))
            changes = []

            # 상태 변경이 삭제 조건이 아니면 일반 변경으로 명시한다. F열이 제외열이어도 이 변경은 표시한다.
            if old_status != new_status:
                changes.append({
                    'col': status_col,
                    'name': _col_name(headers, status_idx),
                    'before': _clean_value(_cell(old_values, status_idx)) or '-',
                    'after': _clean_value(_cell(new_values, status_idx)) or '-'
                })

            for idx in range(max_len):
                col_letter = openpyxl.utils.get_column_letter(idx + 1)
                if col_letter in exclude_cols:
                    continue
                old_val = _clean_value(_cell(old_values, idx))
                new_val = _clean_value(_cell(new_values, idx))
                if old_val != new_val:
                    changes.append({
                        'col': col_letter,
                        'name': _col_name(headers, idx),
                        'before': old_val or '-',
                        'after': new_val or '-'
                    })

            if changes:
                changed.append({
                    'key': key,
                    'old_row_num': old_rec['row_num'],
                    'new_row_num': new_rec['row_num'],
                    'changes': changes
                })

        summary_records = _status_scope_filter(new_records, status_idx, scope)
        summary = _summary_for_records(summary_records, os_col, os_ver_col, physical_col, eosl_col)

        custom_summary = None
        custom_group_col = _clean_value(data.get('custom_group_col', '')).upper()
        if custom_group_col:
            group_idx = col_letter_to_idx(custom_group_col)
            custom_summary = {
                'group_col': custom_group_col,
                'group_label': _col_name(headers, group_idx),
                'counts': _count_by(summary_records, group_idx)
            }

        result = {
            'success': True,
            'options': {
                'key_cols': key_cols,
                'header_row': header_row,
                'data_start_row': data_start_row,
                'status_col': status_col,
                'exclude_cols': sorted(exclude_cols),
                'scope': scope,
                'scope_label': scope_label_map.get(scope, '운영+대기'),
                'os_col': os_col,
                'os_version_col': os_ver_col,
                'physical_col': physical_col,
                'eosl_col': eosl_col,
                'eosl_rule': 'AK열 앞 4자리 기준, 99991231은 미정 처리'
            },
            'compare': {
                'old_count': len(old_records),
                'new_count': len(new_records),
                'added_count': len(added),
                'removed_count': len(removed),
                'status_deleted_count': len(status_deleted),
                'changed_count': len(changed),
                'common_count': len(common_keys)
            },
            # 기존 화면 호환 필드
            'added_count': len(added),
            'removed_count': len(removed),
            'common_count': len(common_keys),
            'added': added,
            'removed': removed,
            'status_deleted': status_deleted,
            'changed': changed,
            'summary': summary,
            'custom_summary': custom_summary,
            'warnings': {
                'old_duplicate_keys': old_data['duplicate_keys'],
                'new_duplicate_keys': new_data['duplicate_keys'],
                'old_empty_key_rows': old_data['empty_key_rows'],
                'new_empty_key_rows': new_data['empty_key_rows']
            }
        }
        result['output_file'] = _make_compare_excel(result)
        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

