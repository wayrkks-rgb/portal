from __future__ import annotations

import os
import re
import shutil
from datetime import datetime

from flask import Blueprint, jsonify, render_template, request, send_file, session
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from application.common import col_letter_to_idx, load_json, require_login, save_json
from application.settings import HISTORY_FILE, MAPPINGS_FILE, OUTPUT_DIR, UPLOAD_DIR

bp = Blueprint("report", __name__)

@bp.route("/report")
@require_login
def report():
    return render_template('main.html', user=session['user'], page='report')

@bp.route("/api/upload-asis", methods=["POST"])
@require_login
def upload_asis():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '파일이 없습니다.'}), 400
    filename = secure_filename(f.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    saved_name = f'asis_{ts}_{filename}'
    path = os.path.join(UPLOAD_DIR, saved_name)
    f.save(path)

    # 파일 분석
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header = rows[0] if rows else []
        preview = rows[1:6] if len(rows) > 1 else []
        total = len(rows) - 1

        # 이력 저장
        history = load_json(HISTORY_FILE)
        history.append({
            'id': ts,
            'type': 'asis',
            'filename': filename,
            'saved_name': saved_name,
            'uploaded_at': datetime.now().isoformat(),
            'uploaded_by': session['user']['name'],
            'row_count': total
        })
        save_json(HISTORY_FILE, history)

        return jsonify({
            'success': True,
            'file_id': ts,
            'saved_name': saved_name,
            'filename': filename,
            'row_count': total,
            'header': list(header),
            'preview': [list(r) for r in preview]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route("/api/upload-tobe", methods=["POST"])
@require_login
def upload_tobe():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '파일이 없습니다.'}), 400
    filename = secure_filename(f.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    saved_name = f'tobe_{ts}_{filename}'
    path = os.path.join(UPLOAD_DIR, saved_name)
    f.save(path)

    try:
        wb = load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        history = load_json(HISTORY_FILE)
        history.append({
            'id': f'tobe_{ts}',
            'type': 'tobe',
            'filename': filename,
            'saved_name': saved_name,
            'uploaded_at': datetime.now().isoformat(),
            'uploaded_by': session['user']['name'],
            'sheet_count': len(sheets)
        })
        save_json(HISTORY_FILE, history)

        return jsonify({
            'success': True,
            'file_id': ts,
            'saved_name': saved_name,
            'filename': filename,
            'sheets': sheets
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route("/api/generate-report", methods=["POST"])
@require_login
def generate_report():
    data = request.get_json()
    asis_name = data.get('asis_file')
    tobe_name = data.get('tobe_file')
    menu_id = data.get('menu_id', 'vm_resource')

    asis_path = os.path.join(UPLOAD_DIR, asis_name)
    tobe_path = os.path.join(UPLOAD_DIR, tobe_name)

    if not os.path.exists(asis_path) or not os.path.exists(tobe_path):
        return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404

    try:
        mappings = load_json(MAPPINGS_FILE)
        menu = next((m for m in mappings['menus'] if m['id'] == menu_id), None)
        if not menu:
            return jsonify({'error': '매핑 설정을 찾을 수 없습니다.'}), 404

        # AS-IS 데이터 로드
        wb_asis = load_workbook(asis_path, read_only=True)
        ws_asis = wb_asis.active
        rows_asis = list(ws_asis.iter_rows(values_only=True))
        wb_asis.close()

        header_row = menu.get('asis_header_row', 1) - 1
        asis_vm_col = col_letter_to_idx(menu['asis_col_vm'])
        asis_cpu_max_col = col_letter_to_idx(menu['asis_col_cpu_max'])
        asis_cpu_avg_col = col_letter_to_idx(menu['asis_col_cpu_avg'])
        asis_mem_max_col = col_letter_to_idx(menu['asis_col_mem_max'])
        asis_mem_avg_col = col_letter_to_idx(menu['asis_col_mem_avg'])

        # VM명 → 데이터 딕셔너리
        vm_data = {}
        for row in rows_asis[header_row + 1:]:
            if row and row[asis_vm_col]:
                vm_name = str(row[asis_vm_col]).strip().lower()
                vm_data[vm_name] = {
                    'cpu_max': row[asis_cpu_max_col],
                    'cpu_avg': row[asis_cpu_avg_col],
                    'mem_max': row[asis_mem_max_col],
                    'mem_avg': row[asis_mem_avg_col],
                }

        # 시트별 VM 매핑 로드
        sheet_vm_map = {}
        for svm in menu.get('sheet_vm_mappings', []):
            sheet = svm.get('sheet_name', '')
            vms = svm.get('vm_list', [])
            sheet_vm_map[sheet] = [v.strip().lower() for v in vms if v.strip()]

        # TO-BE 파일 복사 후 수정
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_filename = f'보고서_{ts}.xlsx'
        out_path = os.path.join(OUTPUT_DIR, out_filename)
        shutil.copy2(tobe_path, out_path)

        wb_tobe = load_workbook(out_path)
        tobe_vm_col = col_letter_to_idx(menu['tobe_col_vm'])
        tobe_cpu_max_col = col_letter_to_idx(menu['tobe_col_cpu_max'])
        tobe_cpu_avg_col = col_letter_to_idx(menu['tobe_col_cpu_avg'])
        tobe_mem_max_col = col_letter_to_idx(menu['tobe_col_mem_max'])
        tobe_mem_avg_col = col_letter_to_idx(menu['tobe_col_mem_avg'])
        tobe_data_start = menu.get('tobe_data_start_row', 7)
        tobe_count_row = menu.get('tobe_count_row', 3)

        results = []
        for sheet_name in wb_tobe.sheetnames:
            ws = wb_tobe[sheet_name]
            matched = 0
            not_found = []

            # 해당 시트에 속한 VM 목록 결정
            allowed_vms = sheet_vm_map.get(sheet_name, None)

            # 데이터 행 처리
            vm_count = 0
            for row_idx in range(tobe_data_start, ws.max_row + 1):
                cell_vm = ws.cell(row=row_idx, column=tobe_vm_col + 1)
                if cell_vm.value is None:
                    continue
                vm_name_tobe = str(cell_vm.value).strip()
                vm_key = vm_name_tobe.lower()

                # 이 시트의 허용 VM 목록 체크
                if allowed_vms is not None and vm_key not in allowed_vms:
                    continue

                vm_count += 1
                if vm_key in vm_data:
                    d = vm_data[vm_key]
                    def fmt(v):
                        if v is None: return None
                        try: return round(float(v), 2)
                        except: return v

                    ws.cell(row=row_idx, column=tobe_cpu_max_col + 1).value = fmt(d['cpu_max'])
                    ws.cell(row=row_idx, column=tobe_cpu_avg_col + 1).value = fmt(d['cpu_avg'])
                    ws.cell(row=row_idx, column=tobe_mem_max_col + 1).value = fmt(d['mem_max'])
                    ws.cell(row=row_idx, column=tobe_mem_avg_col + 1).value = fmt(d['mem_avg'])
                    matched += 1
                else:
                    not_found.append(vm_name_tobe)

            # 3행 서버 대수 업데이트
            count_cell = ws.cell(row=tobe_count_row, column=2)
            if count_cell.value:
                old_text = str(count_cell.value)
                new_text = re.sub(r'(\d+)(개\s*가상서버)', lambda m: f'{vm_count}{m.group(2)}', old_text)
                count_cell.value = new_text

            results.append({
                'sheet': sheet_name,
                'total_in_sheet': vm_count,
                'matched': matched,
                'not_found': not_found
            })

        wb_tobe.save(out_path)

        total_matched = sum(r['matched'] for r in results)
        total_vms = sum(r['total_in_sheet'] for r in results)

        return jsonify({
            'success': True,
            'output_file': out_filename,
            'summary': {
                'total_vms_in_tobe': total_vms,
                'total_matched': total_matched,
                'sheets_processed': len(results)
            },
            'sheet_results': results
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@bp.route("/api/download/<filename>")
@require_login
def download_file(filename):
    path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({'error': '파일 없음'}), 404
    return send_file(path, as_attachment=True, download_name=filename)

