from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..normalization.code_maps import ASSET_STATUS, SERVER_CATEGORY
from ..repositories import AssetRepository
from .integrated_dashboard_service import ACTIVE_STATUS, PHYSICAL_CATEGORY


class AutomatedReportService:
    """Generate database-backed workbooks without manual source Excel uploads."""

    REPORT_TYPES = {"server_status", "physical", "eosl", "resource_usage"}

    def __init__(self, repository: AssetRepository, output_dir: Path) -> None:
        self.repo = repository
        self.output_dir = Path(output_dir)

    def generate(self, report_type: str, start: str | None = None, end: str | None = None) -> Path:
        report_type = report_type.strip().lower()
        if report_type not in self.REPORT_TYPES:
            raise ValueError("지원 보고서는 server_status, physical, eosl, resource_usage입니다.")
        start_day, end_day = self._period(start, end)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.output_dir / f"{report_type}_{end_day.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"
        if report_type == "resource_usage":
            workbook = self._resource_usage(start_day, end_day)
        else:
            snapshot = self.repo.latest_snapshot("ITSM")
            if not snapshot:
                raise ValueError("정상 ITSM 스냅샷이 없습니다.")
            records = [r for r in self.repo.load_itsm_records(int(snapshot["id"])).values() if r.get("status_code") in ACTIVE_STATUS]
            if report_type == "server_status":
                workbook = self._server_status(records, start_day, end_day)
            elif report_type == "physical":
                workbook = self._physical(records, end_day)
            else:
                workbook = self._eosl(records)
        workbook.save(path)
        return path

    def _server_status(self, records: list[dict[str, Any]], start_day: date, end_day: date) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "서버현황"
        category = Counter(SERVER_CATEGORY.get(str(r.get("server_category_code")), "미정") for r in records)
        location = Counter(self._location(r) for r in records)
        status = Counter(ASSET_STATUS.get(str(r.get("status_code")), "미정") for r in records)
        os_count = Counter(str(r.get("os_family") or "미정") for r in records)
        summary = [
            ("기준일", date.today().isoformat()), ("전체", len(records)),
            ("운영", status.get("운영", 0)), ("대기", status.get("대기", 0)),
            ("물리", category.get("물리", 0)), ("논리", category.get("논리", 0)),
            ("IDC", location.get("IDC", 0)), ("DR", location.get("DR", 0)),
        ]
        ws.append(["구분", "대수"])
        for row in summary:
            ws.append(row)
        self._style(ws)

        os_ws = wb.create_sheet("OS별현황")
        os_ws.append(["OS", "대수"])
        for name, count in os_count.most_common():
            os_ws.append([name, count])
        self._style(os_ws)

        detail = wb.create_sheet("자산상세")
        self._write_asset_detail(detail, records)

        change = wb.create_sheet("기간변동")
        change.append(["조회기간", start_day.isoformat(), end_day.isoformat()])
        change.append(["일시", "유형", "자산ID", "서버명", "Hostname", "IP", "위치", "물리/논리", "OS", "변경필드", "이전", "현재"])
        for event in self._change_rows(start_day, end_day):
            change.append(event)
        self._style(change, header_row=2)
        return wb

    def _physical(self, records: list[dict[str, Any]], end_day: date) -> Workbook:
        physical = [r for r in records if r.get("server_category_code") == PHYSICAL_CATEGORY]
        wb = Workbook()
        ws = wb.active
        ws.title = "물리장비현황"
        ws.append(["OS", "IDC", "DR", "합계"])
        grouped: dict[str, Counter[str]] = {}
        for record in physical:
            grouped.setdefault(str(record.get("os_family") or "미정"), Counter())[self._location(record)] += 1
        for os_name, values in sorted(grouped.items()):
            ws.append([os_name, values.get("IDC", 0), values.get("DR", 0), sum(values.values())])
        self._style(ws)

        trend = wb.create_sheet("최근3개월추이")
        trend.append(["월", "물리서버 대수"])
        for month_start in self._last_months(end_day, 3):
            next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            marker = datetime.combine(next_month, time.min).isoformat()
            snapshot = self.repo.conn.execute(
                "SELECT * FROM snapshot WHERE source='ITSM' AND status IN ('SUCCESS','PARTIAL_SUCCESS') AND collected_at<? ORDER BY collected_at DESC,id DESC LIMIT 1",
                (marker,),
            ).fetchone()
            count = 0
            if snapshot:
                count = int(self.repo.conn.execute(
                    "SELECT COUNT(*) AS cnt FROM itsm_asset_snapshot WHERE snapshot_id=? AND status_code IN ('CMSTA010','CMSTA050') AND server_category_code=?",
                    (snapshot["id"], PHYSICAL_CATEGORY),
                ).fetchone()["cnt"])
            trend.append([month_start.strftime("%Y-%m"), count])
        self._style(trend)
        detail = wb.create_sheet("물리장비상세")
        self._write_asset_detail(detail, physical)
        return wb

    def _eosl(self, records: list[dict[str, Any]]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "EOSL현황"
        grouped = Counter(self._eos_bucket(record.get("eos_value")) for record in records)
        ws.append(["구분", "대수"])
        for name in ["종료", "올해 종료", "예정", "미정", "확인필요"]:
            ws.append([name, grouped.get(name, 0)])
        self._style(ws)
        detail = wb.create_sheet("EOSL상세")
        detail.append(["자산ID", "서버명", "Hostname", "물리/논리", "OS", "OS버전", "EOSL", "분류", "위치"])
        for record in records:
            raw = record.get("raw", {})
            detail.append([
                record.get("cm_id"), raw.get("CM_NAME"), record.get("normalized_hostname"),
                SERVER_CATEGORY.get(str(record.get("server_category_code")), "미정"), record.get("os_family"),
                record.get("os_version"), record.get("eos_value"), self._eos_bucket(record.get("eos_value")), self._location(record),
            ])
        self._style(detail)
        return wb

    def _resource_usage(self, start_day: date, end_day: date) -> Workbook:
        rows = self.repo.conn.execute(
            "SELECT * FROM vcenter_resource_daily WHERE stat_date>=? AND stat_date<=? ORDER BY stat_date, entity_type, esxi_host, vm_name",
            (start_day.isoformat(), end_day.isoformat()),
        ).fetchall()
        if not rows:
            raise ValueError("VM_ResourceUsageExport 데이터가 없습니다. 스크립트 연동 후 생성할 수 있습니다.")
        wb = Workbook()
        ws = wb.active
        ws.title = "통합서버자원사용현황"
        ws.append(["일자", "유형", "vCenter", "ESXi", "VM UUID", "VM명", "CPU Max(%)", "CPU Avg(%)", "MEM Max(%)", "MEM Avg(%)", "샘플수"])
        for row in rows:
            ws.append([row["stat_date"], row["entity_type"], row["vcenter_id"], row["esxi_host"], row["vm_uuid"], row["vm_name"], row["cpu_max_pct"], row["cpu_avg_pct"], row["mem_max_pct"], row["mem_avg_pct"], row["sample_count"]])
        self._style(ws)
        return wb

    def _change_rows(self, start_day: date, end_day: date) -> list[list[Any]]:
        events = self.repo.changes(
            "ITSM", 100000,
            datetime.combine(start_day, time.min).isoformat(),
            datetime.combine(end_day + timedelta(days=1), time.min).isoformat(),
        )
        rows: list[list[Any]] = []
        cache: dict[int, dict[str, dict[str, Any]]] = {}
        for event in events:
            if event["event_type"] not in {"ITSM_ASSET_CREATED", "ITSM_RECORD_REMOVED", "ITSM_ASSET_UPDATED", "ITSM_STATUS_TO_UNUSED", "ITSM_STATUS_TO_DISPOSED", "ITSM_ASSET_REACTIVATED"}:
                continue
            current_id = int(event["snapshot_id"])
            cache.setdefault(current_id, self.repo.load_itsm_records(current_id))
            record = cache[current_id].get(event["asset_key"], {})
            if not record and event.get("previous_snapshot_id"):
                previous_id = int(event["previous_snapshot_id"])
                cache.setdefault(previous_id, self.repo.load_itsm_records(previous_id))
                record = cache[previous_id].get(event["asset_key"], {})
            raw = record.get("raw", {})
            rows.append([
                event["detected_at"], event["event_type"], event["asset_key"], raw.get("CM_NAME"),
                record.get("normalized_hostname"), record.get("primary_ip"), self._location(record),
                SERVER_CATEGORY.get(str(record.get("server_category_code")), "미정"), record.get("os_family"),
                event.get("field_name"),
                self._change_value(event.get("field_name"), event.get("event_type"), event.get("old_value")),
                self._change_value(event.get("field_name"), event.get("event_type"), event.get("new_value")),
            ])
        return rows

    @staticmethod
    def _mb_to_gb(value: Any) -> float | None:
        if value in (None, ""):
            return None
        try:
            return round(float(str(value).replace(",", "").strip()) / 1024, 2)
        except (TypeError, ValueError):
            return None

    @classmethod
    def _change_value(cls, field_name: Any, event_type: Any, value: Any) -> Any:
        if value in (None, ""):
            return value
        field = str(field_name or "").upper()
        is_memory = "MEMORY" in field or field.endswith("MEM") or str(event_type or "") == "RV_MEMORY_CHANGED"
        if not is_memory:
            return value
        gb = cls._mb_to_gb(value)
        return f"{gb:g} GB" if gb is not None else value

    @staticmethod
    def _write_asset_detail(ws: Any, records: list[dict[str, Any]]) -> None:
        ws.append(["자산ID", "서버명", "Hostname", "IP", "상태", "물리/논리", "위치", "OS", "OS버전", "CPU Core", "Memory GB", "EOSL", "제조사", "모델"])
        for record in records:
            raw = record.get("raw", {})
            ws.append([
                record.get("cm_id"), raw.get("CM_NAME"), record.get("normalized_hostname"), record.get("primary_ip"),
                ASSET_STATUS.get(str(record.get("status_code")), "미정"), SERVER_CATEGORY.get(str(record.get("server_category_code")), "미정"),
                AutomatedReportService._location(record), record.get("os_family"), record.get("os_version"), record.get("cpu_cores"),
                AutomatedReportService._mb_to_gb(record.get("memory_mb")), record.get("eos_value"), raw.get("CM_MAKE_NAME"), raw.get("CM_MODEL_NAME"),
            ])
        AutomatedReportService._style(ws)

    @staticmethod
    def _style(ws: Any, header_row: int = 1) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[header_row]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = f"A{header_row + 1}"
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
            ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 10)

    @staticmethod
    def _location(record: dict[str, Any]) -> str:
        raw = record.get("raw", {}) or {}
        return "DR" if str(raw.get("CM_DR_YN") or "").upper() in {"Y", "YES", "1"} or str(record.get("environment_code") or "") == "CMOWNCATCD0040" or "DR" in str(raw.get("CM_PLACE") or "").upper() else "IDC"

    @staticmethod
    def _eos_bucket(value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return "확인필요"
        match = re.search(r"(\d{4})", text)
        if not match:
            return "확인필요"
        year = int(match.group(1))
        if year == 9999:
            return "미정"
        if year < date.today().year:
            return "종료"
        if year == date.today().year:
            return "올해 종료"
        return "예정"

    @staticmethod
    def _last_months(end_day: date, count: int) -> list[date]:
        result: list[date] = []
        current = end_day.replace(day=1)
        for _ in range(count):
            result.append(current)
            current = (current - timedelta(days=1)).replace(day=1)
        return list(reversed(result))

    @staticmethod
    def _period(start: str | None, end: str | None) -> tuple[date, date]:
        today = date.today()
        start_day = date.fromisoformat((start or today.replace(day=1).isoformat())[:10])
        end_day = date.fromisoformat((end or today.isoformat())[:10])
        if start_day > end_day:
            raise ValueError("시작일은 종료일보다 늦을 수 없습니다.")
        return start_day, end_day
