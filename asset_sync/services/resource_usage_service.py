from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ..collectors.powercli_resource_collector import PowerCLIResourceUsageCollector
from ..config import AppConfig
from ..repositories import AssetRepository
from ..utils.hashing import canonical_json


class VMResourceUsageExportService:
    """Collect, persist, query and export ESXi/VM resource usage.

    Daily 07:00 processing stores the previous day's usage and links it to the
    vCenter inventory snapshot from the same batch. Arbitrary date ranges are
    calculated from the persisted daily facts; operators do not run PowerCLI from
    the screen.
    """

    COLUMN_ALIASES = {
        "ENTITY_TYPE": "entity_type",
        "TYPE": "entity_type",
        "VCENTER_ID": "vcenter_id",
        "VCENTER": "vcenter_id",
        "SERVICE_NAME": "service_name",
        "SERVICENAME": "service_name",
        "CLUSTER_NAME": "cluster_name",
        "CLUSTERNAME": "cluster_name",
        "ESXI_HOST": "esxi_host",
        "HOST": "esxi_host",
        "HOSTNAME": "esxi_host",
        "VM_UUID": "vm_uuid",
        "VM_NAME": "vm_name",
        "VMNAME": "vm_name",
        "POWER_STATE": "power_state",
        "POWERSTATE": "power_state",
        "ALLOCATED_CPU_CORES": "allocated_cpu_cores",
        "CPUS": "allocated_cpu_cores",
        "ALLOCATED_MEMORY_MB": "allocated_memory_mb",
        "MEMORY_MB": "allocated_memory_mb",
        "CPU_MAX": "cpu_max_pct",
        "CPUMAX": "cpu_max_pct",
        "CPU_MAX_PCT": "cpu_max_pct",
        "CPU_AVG": "cpu_avg_pct",
        "CPUAVG": "cpu_avg_pct",
        "CPU_AVG_PCT": "cpu_avg_pct",
        "MEM_MAX": "mem_max_pct",
        "MEMMAX": "mem_max_pct",
        "MEM_MAX_PCT": "mem_max_pct",
        "MEM_AVG": "mem_avg_pct",
        "MEMAVG": "mem_avg_pct",
        "MEM_AVG_PCT": "mem_avg_pct",
        "SAMPLE_COUNT": "sample_count",
    }

    def __init__(self, config: AppConfig, repository: AssetRepository) -> None:
        self.config = config
        self.repo = repository
        self.settings = config.rvtools.get("resource_usage", {}) or {}

    def daily_status(self) -> dict[str, Any]:
        script = self.config.resolve(
            self.settings.get("script_path", "scripts/collect_vcenter_resource_usage.ps1")
        )
        enabled = bool(self.settings.get("enabled", True))
        if not enabled:
            return {"status": "DISABLED", "message": "통합서버 자원사용률 자동수집이 비활성화되어 있습니다.", "script": script.name}
        if not script.exists():
            return {"status": "PENDING_SCRIPT", "message": "자원사용률 PowerCLI 스크립트가 없습니다.", "script": script.name}
        latest = self.repo.conn.execute(
            "SELECT * FROM resource_usage_run ORDER BY started_at DESC, id DESC LIMIT 1"
        ).fetchone()
        result = {"status": "READY", "message": "07시 자동배치에서 자원사용률을 수집합니다.", "script": script.name}
        if latest:
            result["latest_run"] = dict(latest)
        return result

    def collect_for_batch(
        self,
        daily_batch_id: int,
        vcenter_snapshot_id: int,
        *,
        demo: bool = False,
        period_start: date | None = None,
        period_end: date | None = None,
    ) -> dict[str, Any]:
        end_day = period_end or (date.today() - timedelta(days=1))
        start_day = period_start or end_day
        started_at = datetime.now().isoformat()
        cur = self.repo.conn.execute(
            """
            INSERT INTO resource_usage_run(
                daily_batch_id, vcenter_snapshot_id, period_start, period_end, started_at, status
            ) VALUES (?, ?, ?, ?, ?, 'RUNNING')
            """,
            (daily_batch_id, vcenter_snapshot_id, start_day.isoformat(), end_day.isoformat(), started_at),
        )
        run_id = int(cur.lastrowid)
        try:
            if demo:
                payload = self._demo_payload(vcenter_snapshot_id)
            else:
                if not bool(self.settings.get("enabled", True)):
                    raise RuntimeError("통합서버 자원사용률 자동수집이 비활성화되어 있습니다.")
                payload = PowerCLIResourceUsageCollector(self.config).collect_all(start_day, end_day)
                if payload.get("status") == "FAILED":
                    raise RuntimeError("모든 vCenter 자원사용률 수집이 실패했습니다.")
            hosts, vms = self._enrich_with_inventory(
                payload.get("hosts", []), payload.get("vms", []), vcenter_snapshot_id
            )
            self._replace_run_rows(run_id, end_day.isoformat(), vcenter_snapshot_id, hosts, vms)
            status = str(payload.get("status") or "SUCCESS")
            self.repo.conn.execute(
                """
                UPDATE resource_usage_run
                   SET ended_at=?, status=?, success_scope_json=?, failed_scope_json=?,
                       host_count=?, vm_count=?, metadata_json=?
                 WHERE id=?
                """,
                (
                    datetime.now().isoformat(), status,
                    canonical_json(payload.get("success_scopes", [])),
                    canonical_json(payload.get("failed_scopes", {})),
                    len(hosts), len(vms),
                    canonical_json({"period_start": start_day.isoformat(), "period_end": end_day.isoformat()}),
                    run_id,
                ),
            )
            return {
                "status": status,
                "run_id": run_id,
                "period_start": start_day.isoformat(),
                "period_end": end_day.isoformat(),
                "host_count": len(hosts),
                "vm_count": len(vms),
                "failed_scopes": payload.get("failed_scopes", {}),
            }
        except Exception as exc:
            self.repo.conn.execute(
                "UPDATE resource_usage_run SET ended_at=?, status='FAILED', error_message=? WHERE id=?",
                (datetime.now().isoformat(), str(exc), run_id),
            )
            return {"status": "FAILED", "run_id": run_id, "error": str(exc), "host_count": 0, "vm_count": 0}

    def _demo_payload(self, snapshot_id: int) -> dict[str, Any]:
        records = list(self.repo.load_rv_records(snapshot_id).values())
        vms: list[dict[str, Any]] = []
        host_members: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
        for index, vm in enumerate(records, start=1):
            if vm.get("template_flag") or vm.get("srm_placeholder"):
                continue
            vc = str(vm.get("vcenter") or "DEMO_VCENTER")
            cluster = str(vm.get("cluster_name") or "DEMO_CLUSTER")
            host = str(vm.get("esxi_host") or "DEMO_ESXI")
            row = {
                "vcenter_id": vc,
                "service_name": vc,
                "cluster_name": cluster,
                "esxi_host": host,
                "vm_uuid": vm.get("vm_uuid"),
                "vm_name": vm.get("vm_name"),
                "power_state": vm.get("power_state"),
                "allocated_cpu_cores": vm.get("cpus"),
                "allocated_memory_mb": vm.get("memory_mb"),
                "cpu_max_pct": round(30 + index * 1.7, 2),
                "cpu_avg_pct": round(12 + index * 0.8, 2),
                "mem_max_pct": round(45 + index * 1.3, 2),
                "mem_avg_pct": round(28 + index * 0.7, 2),
                "sample_count": 12,
            }
            vms.append(row)
            host_members[(vc, cluster, host)].append(row)
        hosts = []
        for (vc, cluster, host), members in host_members.items():
            hosts.append({
                "vcenter_id": vc,
                "service_name": vc,
                "cluster_name": cluster,
                "esxi_host": host,
                "allocated_cpu_cores": 64,
                "allocated_memory_mb": 524288,
                "cpu_max_pct": max(float(m["cpu_max_pct"]) for m in members),
                "cpu_avg_pct": round(sum(float(m["cpu_avg_pct"]) for m in members) / len(members), 2),
                "mem_max_pct": max(float(m["mem_max_pct"]) for m in members),
                "mem_avg_pct": round(sum(float(m["mem_avg_pct"]) for m in members) / len(members), 2),
                "sample_count": 12,
            })
        return {"status": "SUCCESS", "hosts": hosts, "vms": vms, "success_scopes": sorted({r["vcenter_id"] for r in vms}), "failed_scopes": {}}

    def _enrich_with_inventory(
        self,
        host_rows: list[dict[str, Any]],
        usage_rows: list[dict[str, Any]],
        snapshot_id: int,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        inventory = [r for r in self.repo.load_rv_records(snapshot_id).values() if not r.get("template_flag") and not r.get("srm_placeholder")]
        by_uuid = {(str(r.get("vcenter") or ""), str(r.get("vm_uuid") or "").lower()): r for r in inventory if r.get("vm_uuid")}
        by_name = {(str(r.get("vcenter") or ""), str(r.get("vm_name") or "").lower()): r for r in inventory if r.get("vm_name")}
        usage_map: dict[str, dict[str, Any]] = {}
        for raw in usage_rows:
            row = self._normalize(raw)
            vc = str(row.get("vcenter_id") or "")
            match = None
            if row.get("vm_uuid"):
                match = by_uuid.get((vc, str(row["vm_uuid"]).lower()))
            if not match and row.get("vm_name"):
                match = by_name.get((vc, str(row["vm_name"]).lower()))
            if match:
                row.update({
                    "asset_key": match.get("asset_key"),
                    "vcenter_id": match.get("vcenter") or vc,
                    "cluster_name": match.get("cluster_name"),
                    "esxi_host": match.get("esxi_host"),
                    "vm_uuid": match.get("vm_uuid") or row.get("vm_uuid"),
                    "vm_name": match.get("vm_name") or row.get("vm_name"),
                    "power_state": match.get("power_state"),
                    "allocated_cpu_cores": match.get("cpus"),
                    "allocated_memory_mb": match.get("memory_mb"),
                    "inventory_status": "CURRENT",
                })
            key = str(row.get("asset_key") or f"{row.get('vcenter_id')}|{row.get('vm_uuid') or row.get('vm_name')}")
            usage_map[key] = row

        vc_names = {
            str(item.get("id") or item.get("name") or ""): str(item.get("name") or item.get("id") or "")
            for item in self.config.rvtools.get("vcenters", [])
        }
        final_vms: list[dict[str, Any]] = []
        for vm in inventory:
            key = str(vm.get("asset_key"))
            row = usage_map.pop(key, None) or {
                "vcenter_id": vm.get("vcenter"),
                "vm_uuid": vm.get("vm_uuid"),
                "vm_name": vm.get("vm_name"),
                "collection_status": "NO_STAT",
                "sample_count": 0,
            }
            row.update({
                "asset_key": vm.get("asset_key"),
                "vcenter_id": vm.get("vcenter"),
                "service_name": row.get("service_name") or vc_names.get(str(vm.get("vcenter") or "")) or vm.get("vcenter"),
                "cluster_name": vm.get("cluster_name"),
                "esxi_host": vm.get("esxi_host"),
                "vm_uuid": vm.get("vm_uuid"),
                "vm_name": vm.get("vm_name"),
                "power_state": vm.get("power_state"),
                "allocated_cpu_cores": vm.get("cpus"),
                "allocated_memory_mb": vm.get("memory_mb"),
                "inventory_status": "CURRENT",
            })
            final_vms.append(row)
        for orphan in usage_map.values():
            orphan["inventory_status"] = "NOT_IN_CURRENT_INVENTORY"
            final_vms.append(orphan)

        host_vm_count: dict[tuple[str, str], int] = defaultdict(int)
        for vm in final_vms:
            if vm.get("inventory_status") == "CURRENT" and vm.get("esxi_host"):
                host_vm_count[(str(vm.get("vcenter_id") or ""), str(vm.get("esxi_host")))] += 1
        normalized_hosts = []
        for raw in host_rows:
            row = self._normalize(raw)
            row["entity_type"] = "ESXI"
            row["service_name"] = row.get("service_name") or vc_names.get(str(row.get("vcenter_id") or "")) or row.get("vcenter_id")
            row["vm_count"] = host_vm_count.get((str(row.get("vcenter_id") or ""), str(row.get("esxi_host") or "")), 0)
            normalized_hosts.append(row)
        known_hosts = {(str(r.get("vcenter_id") or ""), str(r.get("esxi_host") or "")) for r in normalized_hosts}
        for (vc, host), vm_count in host_vm_count.items():
            if (vc, host) not in known_hosts:
                sample_vm = next((v for v in final_vms if str(v.get("vcenter_id") or "") == vc and str(v.get("esxi_host") or "") == host), {})
                normalized_hosts.append({
                    "entity_type": "ESXI", "vcenter_id": vc, "service_name": vc_names.get(vc) or vc,
                    "cluster_name": sample_vm.get("cluster_name"), "esxi_host": host, "vm_count": vm_count,
                    "collection_status": "NO_STAT", "sample_count": 0,
                })
        return normalized_hosts, final_vms

    def _replace_run_rows(
        self,
        run_id: int,
        stat_date: str,
        snapshot_id: int,
        hosts: list[dict[str, Any]],
        vms: list[dict[str, Any]],
    ) -> None:
        now = datetime.now().isoformat()
        self.repo.conn.execute("DELETE FROM host_resource_usage_daily WHERE run_id=?", (run_id,))
        self.repo.conn.execute("DELETE FROM vm_resource_usage_daily WHERE run_id=?", (run_id,))
        self.repo.conn.executemany(
            """
            INSERT INTO host_resource_usage_daily(
                run_id, stat_date, vcenter_id, service_name, cluster_name, esxi_host, vm_count,
                allocated_cpu_cores, allocated_memory_mb, cpu_max_pct, cpu_avg_pct, mem_max_pct,
                mem_avg_pct, sample_count, collection_status, raw_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [(
                run_id, stat_date, str(r.get("vcenter_id") or "UNKNOWN"), r.get("service_name"),
                r.get("cluster_name"), str(r.get("esxi_host") or "UNKNOWN"), int(r.get("vm_count") or 0),
                self._int(r.get("allocated_cpu_cores")), self._int(r.get("allocated_memory_mb")),
                self._float(r.get("cpu_max_pct")), self._float(r.get("cpu_avg_pct")),
                self._float(r.get("mem_max_pct")), self._float(r.get("mem_avg_pct")),
                self._int(r.get("sample_count")) or 0, r.get("collection_status", "SUCCESS"),
                canonical_json(r.get("raw", r)), now,
            ) for r in hosts],
        )
        self.repo.conn.executemany(
            """
            INSERT INTO vm_resource_usage_daily(
                run_id, stat_date, vcenter_snapshot_id, asset_key, vcenter_id, service_name,
                cluster_name, esxi_host, vm_uuid, vm_name, power_state, allocated_cpu_cores,
                allocated_memory_mb, cpu_max_pct, cpu_avg_pct, mem_max_pct, mem_avg_pct,
                sample_count, inventory_status, collection_status, raw_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [(
                run_id, stat_date, snapshot_id, r.get("asset_key"), str(r.get("vcenter_id") or "UNKNOWN"),
                r.get("service_name"), r.get("cluster_name"), r.get("esxi_host"), r.get("vm_uuid"),
                str(r.get("vm_name") or "UNKNOWN"), r.get("power_state"), self._int(r.get("allocated_cpu_cores")),
                self._int(r.get("allocated_memory_mb")), self._float(r.get("cpu_max_pct")),
                self._float(r.get("cpu_avg_pct")), self._float(r.get("mem_max_pct")),
                self._float(r.get("mem_avg_pct")), self._int(r.get("sample_count")) or 0,
                r.get("inventory_status", "CURRENT"), r.get("collection_status", "SUCCESS"),
                canonical_json(r.get("raw", r)), now,
            ) for r in vms],
        )

    def summary(
        self,
        start: str,
        end: str,
        *,
        vcenter_id: str | None = None,
        cluster_name: str | None = None,
        esxi_host: str | None = None,
    ) -> dict[str, Any]:
        start_day = date.fromisoformat(start[:10])
        end_day = date.fromisoformat(end[:10])
        if start_day > end_day:
            raise ValueError("시작일은 종료일보다 늦을 수 없습니다.")
        filters = ["stat_date>=?", "stat_date<=?"]
        params: list[Any] = [start_day.isoformat(), end_day.isoformat()]
        for column, value in (("vcenter_id", vcenter_id), ("cluster_name", cluster_name), ("esxi_host", esxi_host)):
            if value:
                filters.append(f"{column}=?")
                params.append(value)
        where = " AND ".join(filters)
        host_rows = [dict(r) for r in self.repo.conn.execute(
            f"SELECT * FROM host_resource_usage_daily WHERE {where} ORDER BY stat_date, service_name, esxi_host", params
        ).fetchall()]
        vm_rows = [dict(r) for r in self.repo.conn.execute(
            f"SELECT * FROM vm_resource_usage_daily WHERE {where} ORDER BY stat_date, service_name, esxi_host, vm_name", params
        ).fetchall()]
        hosts = self._aggregate(host_rows, ["vcenter_id", "service_name", "cluster_name", "esxi_host"], host=True)
        vms = self._aggregate(vm_rows, ["vcenter_id", "service_name", "vm_uuid", "vm_name"], host=False)
        changes = self._vm_configuration_changes(start_day, end_day, vcenter_id, esxi_host)
        return {
            "period": {"start": start_day.isoformat(), "end": end_day.isoformat()},
            "hosts": hosts,
            "vms": vms,
            "changes": changes,
            "filters": self._available_filters(),
            "summary": {
                "host_count": len(hosts), "vm_count": len(vms),
                "cpu_changed": sum(1 for r in changes if r["event_type"] == "RV_CPU_CHANGED"),
                "memory_changed": sum(1 for r in changes if r["event_type"] == "RV_MEMORY_CHANGED"),
                "vm_added": sum(1 for r in changes if r["event_type"] == "RV_NEW"),
                "vm_removed": sum(1 for r in changes if r["event_type"] == "RV_REMOVED"),
            },
        }

    def export_xlsx(self, start: str, end: str, target_dir: Path, **filters: Any) -> Path:
        data = self.summary(start, end, **filters)
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"통합서버_자원사용현황_{start}_{end}.xlsx"
        wb = Workbook()
        ws_host = wb.active
        ws_host.title = "HostResourceUsage"
        self._write_sheet(ws_host, [
            ("서비스명", "service_name"), ("vCenter", "vcenter_id"), ("Cluster", "cluster_name"),
            ("통합기", "esxi_host"), ("VM 대수", "vm_count"), ("실제 CPU Core", "allocated_cpu_cores"),
            ("실제 Memory GB", "allocated_memory_gb"), ("CPU MAX %", "cpu_max_pct"),
            ("CPU AVG %", "cpu_avg_pct"), ("MEM MAX %", "mem_max_pct"), ("MEM AVG %", "mem_avg_pct"),
        ], data["hosts"])
        ws_vm = wb.create_sheet("VMsResource")
        self._write_sheet(ws_vm, [
            ("서비스명", "service_name"), ("vCenter", "vcenter_id"), ("Cluster", "cluster_name"),
            ("통합기", "esxi_host"), ("VM UUID", "vm_uuid"), ("VM명", "vm_name"),
            ("전원상태", "power_state"), ("실제 CPU Core", "allocated_cpu_cores"),
            ("실제 Memory GB", "allocated_memory_gb"), ("CPU MAX %", "cpu_max_pct"),
            ("CPU AVG %", "cpu_avg_pct"), ("MEM MAX %", "mem_max_pct"), ("MEM AVG %", "mem_avg_pct"),
        ], data["vms"])
        ws_change = wb.create_sheet("VMChangeHistory")
        self._write_sheet(ws_change, [
            ("변경일시", "detected_at"), ("vCenter", "vcenter_id"), ("통합기", "esxi_host"),
            ("VM명", "vm_name"), ("변경유형", "event_type"), ("변경필드", "field_name"),
            ("이전값", "old_value_display"), ("현재값", "new_value_display"),
        ], self._change_export_rows(data["changes"]))
        wb.save(target)
        wb.close()
        return target

    def import_file(self, file_path: Path, stat_date: str | None = None) -> dict[str, Any]:
        """Compatibility import for previously exported JSON/CSV/XLSX files."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(path)
        target_date = stat_date or (date.today() - timedelta(days=1)).isoformat()
        suffix = path.suffix.lower()
        if suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8-sig"))
            rows = data if isinstance(data, list) else data.get("records", [])
        elif suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
        elif suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(v or "").strip() for v in next(values)]
            rows = [dict(zip(headers, row)) for row in values]
            workbook.close()
        else:
            raise ValueError("지원 파일은 JSON, CSV, XLSX입니다.")
        normalized = [self._normalize(row) for row in rows if any(v not in (None, "") for v in row.values())]
        count = self.repo.replace_resource_usage(target_date, normalized)
        return {"status": "SUCCESS", "stat_date": target_date, "count": count, "source": path.name}

    def _aggregate(self, rows: list[dict[str, Any]], keys: list[str], *, host: bool) -> list[dict[str, Any]]:
        grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            grouped[tuple(row.get(k) for k in keys)].append(row)
        result: list[dict[str, Any]] = []
        for key, group in grouped.items():
            latest = sorted(group, key=lambda r: (r.get("stat_date") or "", r.get("id") or 0))[-1]
            item = {name: value for name, value in zip(keys, key)}
            item.update({
                "cluster_name": latest.get("cluster_name"),
                "esxi_host": latest.get("esxi_host"),
                "vm_count": latest.get("vm_count") if host else None,
                "power_state": latest.get("power_state"),
                "allocated_cpu_cores": latest.get("allocated_cpu_cores"),
                "allocated_memory_mb": latest.get("allocated_memory_mb"),
                "allocated_memory_gb": self._mb_to_gb(latest.get("allocated_memory_mb")),
                "cpu_max_pct": self._max(group, "cpu_max_pct"),
                "cpu_avg_pct": self._weighted_avg(group, "cpu_avg_pct"),
                "mem_max_pct": self._max(group, "mem_max_pct"),
                "mem_avg_pct": self._weighted_avg(group, "mem_avg_pct"),
                "sample_count": sum(int(r.get("sample_count") or 0) for r in group),
                "latest_stat_date": latest.get("stat_date"),
            })
            result.append(item)
        result.sort(key=lambda r: tuple(str(r.get(k) or "") for k in keys))
        return result

    def _vm_configuration_changes(self, start_day: date, end_day: date, vcenter_id: str | None, esxi_host: str | None) -> list[dict[str, Any]]:
        start_dt = datetime.combine(start_day, datetime.min.time()).isoformat()
        end_dt = datetime.combine(end_day + timedelta(days=1), datetime.min.time()).isoformat()
        events = self.repo.changes("RVTOOLS", 100000, start_dt, end_dt)
        wanted = {"RV_NEW", "RV_REMOVED", "RV_CPU_CHANGED", "RV_MEMORY_CHANGED", "RV_HOST_CHANGED", "RV_VCENTER_CHANGED"}
        cache: dict[int, dict[str, dict[str, Any]]] = {}
        result = []
        for event in events:
            if event.get("event_type") not in wanted:
                continue
            records = []
            for sid in (event.get("snapshot_id"), event.get("previous_snapshot_id")):
                if not sid:
                    continue
                sid = int(sid)
                cache.setdefault(sid, self.repo.load_rv_records(sid))
                if cache[sid].get(event["asset_key"]):
                    records.append(cache[sid][event["asset_key"]])
            vm = records[0] if records else {}
            vc = str(vm.get("vcenter") or "")
            host = str(vm.get("esxi_host") or "")
            if vcenter_id and vc != vcenter_id:
                continue
            if esxi_host and host != esxi_host and event.get("old_value") != esxi_host and event.get("new_value") != esxi_host:
                continue
            result.append({
                "detected_at": event.get("detected_at"), "asset_key": event.get("asset_key"),
                "vcenter_id": vc, "esxi_host": host, "vm_name": vm.get("vm_name") or event.get("asset_key"),
                "event_type": event.get("event_type"), "field_name": event.get("field_name"),
                "old_value": event.get("old_value"), "new_value": event.get("new_value"),
            })
        return result

    def _distinct(self, column: str, where: str = "") -> list[str]:
        # 컬럼명으로 읽는다. MySQL 커서는 dict 를 돌려주므로 위치 색인은 쓸 수 없다.
        rows = self.repo.conn.execute(
            f"SELECT DISTINCT {column} AS value FROM host_resource_usage_daily {where} ORDER BY {column}"
        ).fetchall()
        return [str(row["value"]) for row in rows if row["value"]]

    def _available_filters(self) -> dict[str, list[str]]:
        return {
            "vcenters": self._distinct("vcenter_id"),
            "clusters": self._distinct("cluster_name", "WHERE cluster_name IS NOT NULL"),
            "hosts": self._distinct("esxi_host"),
        }

    def _normalize(self, raw: dict[str, Any]) -> dict[str, Any]:
        mapped: dict[str, Any] = {}
        for key, value in raw.items():
            alias = self.COLUMN_ALIASES.get(str(key).strip().upper())
            if alias:
                mapped[alias] = value
        entity_type = str(mapped.get("entity_type") or ("VM" if mapped.get("vm_name") or mapped.get("vm_uuid") else "ESXI")).upper()
        if entity_type not in {"VM", "ESXI"}:
            raise ValueError(f"ENTITY_TYPE은 VM 또는 ESXI여야 합니다: {entity_type}")
        return {
            "entity_type": entity_type,
            "vcenter_id": self._text(mapped.get("vcenter_id")), "service_name": self._text(mapped.get("service_name")),
            "cluster_name": self._text(mapped.get("cluster_name")), "esxi_host": self._text(mapped.get("esxi_host")),
            "vm_uuid": self._text(mapped.get("vm_uuid")), "vm_name": self._text(mapped.get("vm_name")),
            "power_state": self._text(mapped.get("power_state")),
            "allocated_cpu_cores": self._int(mapped.get("allocated_cpu_cores")),
            "allocated_memory_mb": self._int(mapped.get("allocated_memory_mb")),
            "cpu_max_pct": self._float(mapped.get("cpu_max_pct")), "cpu_avg_pct": self._float(mapped.get("cpu_avg_pct")),
            "mem_max_pct": self._float(mapped.get("mem_max_pct")), "mem_avg_pct": self._float(mapped.get("mem_avg_pct")),
            "sample_count": self._int(mapped.get("sample_count")) or 0,
            "collection_status": "SUCCESS", "source_name": "VM_ResourceUsageExport", "raw": raw,
        }

    @classmethod
    def _change_export_rows(cls, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for source in rows:
            row = dict(source)
            is_memory = row.get("event_type") == "RV_MEMORY_CHANGED" or "MEMORY" in str(row.get("field_name") or "").upper()
            if is_memory:
                row["old_value_display"] = cls._memory_change_display(row.get("old_value"))
                row["new_value_display"] = cls._memory_change_display(row.get("new_value"))
            else:
                row["old_value_display"] = row.get("old_value")
                row["new_value_display"] = row.get("new_value")
            result.append(row)
        return result

    @classmethod
    def _memory_change_display(cls, value: Any) -> str | None:
        if value in (None, ""):
            return None
        try:
            gb = cls._mb_to_gb(value)
        except (TypeError, ValueError):
            return str(value)
        if gb is None:
            return None
        return f"{gb:g} GB"

    @staticmethod
    def _mb_to_gb(value: Any) -> float | None:
        if value in (None, ""):
            return None
        return round(float(str(value).replace(",", "").strip()) / 1024, 2)

    @staticmethod
    def _write_sheet(ws: Any, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
        ws.append([label for label, _ in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="ED7D31")
        for row in rows:
            ws.append([row.get(key) for _, key in columns])
        for index, (_, key) in enumerate(columns, start=1):
            if key.endswith("_gb"):
                for cell in ws.iter_cols(min_col=index, max_col=index, min_row=2):
                    for item in cell:
                        item.number_format = '0.##'
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
            ws.column_dimensions[column[0].column_letter].width = width

    @staticmethod
    def _max(rows: list[dict[str, Any]], field: str) -> float | None:
        values = [float(r[field]) for r in rows if r.get(field) is not None]
        return round(max(values), 2) if values else None

    @staticmethod
    def _weighted_avg(rows: list[dict[str, Any]], field: str) -> float | None:
        values = [(float(r[field]), int(r.get("sample_count") or 0)) for r in rows if r.get(field) is not None]
        if not values:
            return None
        weight = sum(w for _, w in values)
        return round(sum(v * (w or 1) for v, w in values) / (weight or len(values)), 2)

    @staticmethod
    def _text(value: Any) -> str | None:
        text = str(value or "").strip()
        return text or None

    @staticmethod
    def _float(value: Any) -> float | None:
        if value in (None, ""):
            return None
        return float(str(value).replace("%", "").strip())

    @staticmethod
    def _int(value: Any) -> int | None:
        if value in (None, ""):
            return None
        return int(float(str(value).replace(",", "").strip()))
