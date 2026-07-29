from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import AppConfig

LOGGER = logging.getLogger(__name__)

_REQUIRED = {"VM", "Powerstate", "CPUs", "Memory"}
_OPTIONAL = [
    "Template", "SRM Placeholder", "DNS Name", "Primary IP Address",
    *[f"Network #{i}" for i in range(1, 9)],
    "OS according to the configuration file", "OS according to the VMware Tools", "Datacenter",
    "Cluster", "Host", "VM ID", "SMBIOS UUID", "VM UUID", "VI SDK Server",
    "Creation date", "Change Version",
]


class RVToolsCollectionError(RuntimeError):
    pass


class RVToolsFileCollector:
    """Load and combine RVTools vInfo worksheets from one or more vCenters."""

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def collect(self, files: list[Path] | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        incoming = self.config.resolve(self.config.rvtools.get("incoming_dir", "data/incoming/rvtools"))
        files = files or sorted(incoming.glob("*.xlsx"))
        if not files:
            raise RVToolsCollectionError(f"RVTools 파일이 없습니다: {incoming}")

        records: list[dict[str, Any]] = []
        success_scopes: list[str] = []
        failed: dict[str, str] = {}
        failed_files: list[dict[str, str]] = []
        file_meta: list[dict[str, Any]] = []
        minimum_size = int(self.config.rvtools.get("minimum_file_size_bytes", 1024))
        sheet_name = str(self.config.rvtools.get("sheet_name", "vInfo"))

        for path in files:
            try:
                if path.stat().st_size < minimum_size:
                    raise RVToolsCollectionError("파일 크기가 최소 기준보다 작습니다.")
                workbook = load_workbook(path, read_only=True, data_only=True)
                if sheet_name not in workbook.sheetnames:
                    raise RVToolsCollectionError(f"{sheet_name} 시트가 없습니다.")
                ws = workbook[sheet_name]
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(v).strip() if v is not None else "" for v in header_row]
                header_set = set(headers)
                missing = sorted(_REQUIRED - header_set)
                if missing:
                    raise RVToolsCollectionError(f"필수 컬럼 누락: {', '.join(missing)}")
                index = {name: i for i, name in enumerate(headers) if name}
                count = 0
                inferred_scope = path.stem
                for row in ws.iter_rows(min_row=2, values_only=True):
                    values = list(row)
                    if not any(v is not None and str(v).strip() for v in values):
                        continue
                    item = {name: values[idx] if idx < len(values) else None for name, idx in index.items()}
                    scope = str(item.get("VI SDK Server") or inferred_scope).strip()
                    item["_source_file"] = str(path)
                    item["_vcenter_scope"] = scope
                    records.append(item)
                    count += 1
                workbook.close()
                if count == 0:
                    raise RVToolsCollectionError("vInfo 데이터 행이 0건입니다.")
                scope_name = next((str(r.get("_vcenter_scope")) for r in records[-count:] if r.get("_vcenter_scope")), inferred_scope)
                success_scopes.append(scope_name)
                file_meta.append({"file": str(path), "rows": count, "scope": scope_name, "optional_columns": [c for c in _OPTIONAL if c in header_set]})
            except Exception as exc:
                LOGGER.exception("RVTools file validation failed: %s", path)
                failed[path.stem] = str(exc)
                failed_files.append({"file": str(path), "error": str(exc)})

        if not records:
            raise RVToolsCollectionError("정상 처리된 RVTools 파일이 없습니다.")
        return records, {"success_scopes": sorted(set(success_scopes)), "failed_scopes": failed, "files": file_meta, "failed_files": failed_files}
