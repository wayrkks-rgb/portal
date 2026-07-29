from __future__ import annotations

import json
import logging
import os
import shutil
import socket
import subprocess
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ..config import AppConfig

LOGGER = logging.getLogger(__name__)

_REQUIRED = {"VM", "Powerstate", "CPUs", "Memory"}


class PowerCLICollectionError(RuntimeError):
    pass


class PowerCLICollector:
    """Collect vCenter VM inventory through PowerCLI.

    One PowerShell process is executed per enabled vCenter. Credentials are passed
    through the child-process environment, never command-line arguments. The
    PowerShell script returns JSON; this class optionally stores a daily XLSX copy
    for audit/inspection while SQLite remains the comparison source of truth.
    """

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def enabled_vcenters(self) -> list[dict[str, Any]]:
        return [item for item in self.config.rvtools.get("vcenters", []) if bool(item.get("enabled", True))]

    @staticmethod
    def test_network(entry: dict[str, Any], timeout: float = 5.0) -> dict[str, Any]:
        server = str(entry.get("server") or "").strip()
        port = int(entry.get("port") or 443)
        if not server:
            return {"status": "FAILED", "server": "", "port": port, "error": "vCenter 주소가 없습니다."}
        try:
            with socket.create_connection((server, port), timeout=timeout):
                return {"status": "SUCCESS", "server": server, "port": port}
        except OSError as exc:
            return {"status": "FAILED", "server": server, "port": port, "error": str(exc)}

    def _effective_entry(self, entry: dict[str, Any]) -> dict[str, Any]:
        resolved = dict(entry)
        if entry.get("auth_profile") is None and any(key in entry for key in ("auth_mode", "username", "password", "bypass_ssl_check")):
            profile = "CUSTOM"
        else:
            profile = str(entry.get("auth_profile", "COMMON")).upper()
        resolved["auth_profile"] = profile
        resolved["port"] = int(entry.get("port") or self.config.rvtools.get("default_port", 443))
        if profile == "COMMON":
            resolved["auth_mode"] = str(self.config.rvtools.get("default_auth_mode", "CREDENTIAL")).upper()
            resolved["username"] = str(self.config.rvtools.get("default_username", ""))
            resolved["password"] = str(self.config.rvtools.get("default_password", ""))
            resolved["bypass_ssl_check"] = bool(self.config.rvtools.get("default_bypass_ssl_check", False))
        else:
            resolved["auth_mode"] = str(entry.get("auth_mode", "CREDENTIAL")).upper()
            resolved["bypass_ssl_check"] = bool(entry.get("bypass_ssl_check", False))
        return resolved

    def _resolve_executable(self) -> str | None:
        configured = str(self.config.rvtools.get("powershell_path") or "powershell.exe").strip()
        path = Path(configured)
        if path.is_absolute() and path.exists():
            return str(path)
        return shutil.which(configured)

    def _resolve_script(self) -> Path:
        return self.config.resolve(self.config.rvtools.get("script_path", "scripts/collect_vcenter_inventory.ps1"))

    def _build_environment(self, entry: dict[str, Any], output_json: Path) -> dict[str, str]:
        resolved = self._effective_entry(entry)
        server = str(resolved.get("server") or "").strip()
        if not server:
            raise PowerCLICollectionError("vCenter 주소가 없습니다.")
        auth_mode = str(resolved.get("auth_mode") or "CREDENTIAL").upper()
        if auth_mode not in {"CREDENTIAL", "PASS_THROUGH"}:
            raise PowerCLICollectionError(f"지원하지 않는 인증방식입니다: {auth_mode}")
        username = str(resolved.get("username") or "")
        password = str(resolved.get("password") or "")
        if auth_mode == "CREDENTIAL" and (not username.strip() or not password):
            raise PowerCLICollectionError("계정 인증 방식에는 vCenter 사용자명과 비밀번호가 필요합니다.")

        env = dict(os.environ)
        env.update(
            {
                "VCENTER_SERVER": server,
                "VCENTER_PORT": str(resolved.get("port") or 443),
                "VCENTER_ID": str(resolved.get("id") or resolved.get("name") or "UNKNOWN"),
                "VCENTER_NAME": str(resolved.get("name") or resolved.get("id") or "UNKNOWN"),
                "VCENTER_AUTH_MODE": auth_mode,
                "VCENTER_USERNAME": username,
                "VCENTER_PASSWORD": password,
                "VCENTER_IGNORE_CERT": "true" if bool(resolved.get("bypass_ssl_check", False)) else "false",
                "VCENTER_OUTPUT_JSON": str(output_json),
            }
        )
        return env

    def _load_records(self, path: Path) -> list[dict[str, Any]]:
        if not path.exists():
            raise PowerCLICollectionError(f"PowerCLI 결과 JSON이 생성되지 않았습니다: {path}")
        raw = path.read_text(encoding="utf-8-sig").strip()
        if not raw:
            raise PowerCLICollectionError("PowerCLI 결과 JSON이 비어 있습니다.")
        data = json.loads(raw)
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            raise PowerCLICollectionError("PowerCLI 결과 JSON 최상위 값은 배열이어야 합니다.")
        records = [dict(item) for item in data if isinstance(item, dict)]
        if not records:
            raise PowerCLICollectionError("PowerCLI에서 조회된 VM이 없습니다.")
        missing = sorted(_REQUIRED - set(records[0]))
        if missing:
            raise PowerCLICollectionError(f"PowerCLI 필수 필드 누락: {', '.join(missing)}")
        return records

    def _write_xlsx(self, records: list[dict[str, Any]], target: Path) -> Path:
        target.parent.mkdir(parents=True, exist_ok=True)
        headers: list[str] = []
        seen: set[str] = set()
        for row in records:
            for key in row:
                if key not in seen:
                    headers.append(key)
                    seen.add(key)
        wb = Workbook()
        ws = wb.active
        ws.title = "vInfo"
        ws.append(headers)
        for row in records:
            ws.append([row.get(key) for key in headers])
        wb.save(target)
        wb.close()
        return target

    @staticmethod
    def _classify_failure_stage(message: str) -> str:
        lowered = message.lower()
        if "vmware.vimautomation.core" in lowered or "import-module" in lowered or "module" in lowered and "not" in lowered:
            return "POWERCLI_MODULE"
        if "connect-viserver" in lowered or "authentication" in lowered or "credential" in lowered or "login" in lowered:
            return "VCENTER_AUTH"
        if "certificate" in lowered or "ssl" in lowered or "tls" in lowered:
            return "CERTIFICATE"
        if "permission" in lowered or "privilege" in lowered or "not authorized" in lowered:
            return "VCENTER_PERMISSION"
        return "POWERCLI"

    def run_one(self, entry: dict[str, Any], *, test_run: bool = False) -> dict[str, Any]:
        resolved = self._effective_entry(entry)
        vc_id = str(resolved.get("id") or resolved.get("name") or "UNKNOWN")
        name = str(resolved.get("name") or vc_id)
        network = self.test_network(resolved)
        if network["status"] != "SUCCESS":
            return {"id": vc_id, "name": name, "status": "FAILED", "stage": "NETWORK", "network": network, "error": network.get("error")}

        executable = self._resolve_executable()
        if not executable:
            return {"id": vc_id, "name": name, "status": "FAILED", "stage": "POWERSHELL", "network": network, "error": "PowerShell 실행파일을 찾을 수 없습니다."}
        script = self._resolve_script()
        if not script.exists():
            return {"id": vc_id, "name": name, "status": "FAILED", "stage": "SCRIPT", "network": network, "error": f"PowerCLI 수집 스크립트가 없습니다: {script}"}

        temp_dir = self.config.resolve(self.config.rvtools.get("temp_dir", "data/temp/powercli"))
        temp_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        output_json = temp_dir / f"{'TEST_' if test_run else ''}{vc_id}_{stamp}.json"
        try:
            env = self._build_environment(resolved, output_json)
        except PowerCLICollectionError as exc:
            return {"id": vc_id, "name": name, "status": "FAILED", "stage": "CONFIG", "network": network, "error": str(exc)}

        command = [
            executable,
            "-NoLogo",
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-OutputPath",
            str(output_json),
        ]
        timeout = int(resolved.get("timeout_seconds") or self.config.rvtools.get("timeout_seconds", 1800))
        retry_count = int(resolved.get("retry_count") or self.config.rvtools.get("retry_count", 1))
        log_scope = vc_id if not self.config.security.get("display_vcenter_server_in_logs", False) else str(resolved.get("server", vc_id))
        LOGGER.info("PowerCLI collection start: vcenter=%s auth_profile=%s auth_mode=%s", log_scope, resolved.get("auth_profile"), resolved.get("auth_mode"))

        last_error = ""
        proc: subprocess.CompletedProcess[str] | None = None
        for attempt in range(retry_count + 1):
            try:
                proc = subprocess.run(
                    command,
                    shell=False,
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                    cwd=str(self.config.root_dir),
                    env=env,
                )
                if proc.returncode == 0:
                    records = self._load_records(output_json)
                    snapshot_path: Path | None = None
                    if bool(self.config.rvtools.get("export_xlsx", True)):
                        if test_run:
                            snapshot_dir = self.config.resolve("data/test/powercli")
                        else:
                            date_dir = datetime.now().strftime("%Y%m%d")
                            snapshot_dir = self.config.resolve(self.config.rvtools.get("snapshot_dir", "data/archive/vcenter")) / date_dir
                        snapshot_path = self._write_xlsx(records, snapshot_dir / f"vcenter_{vc_id}_{stamp}.xlsx")
                    return {
                        "id": vc_id,
                        "name": name,
                        "status": "SUCCESS",
                        "stage": "VALIDATED",
                        "network": network,
                        "records": records,
                        "row_count": len(records),
                        "json_file": str(output_json),
                        "xlsx_file": str(snapshot_path) if snapshot_path else None,
                        "attempt": attempt + 1,
                        "stdout": (proc.stdout or "")[-2000:],
                    }
                last_error = (proc.stderr or proc.stdout or "PowerCLI 실행에 실패했습니다.")[-4000:]
            except subprocess.TimeoutExpired:
                last_error = "PowerCLI 실행시간 초과"
            except (OSError, ValueError, json.JSONDecodeError, PowerCLICollectionError) as exc:
                last_error = str(exc)
            if attempt < retry_count:
                time.sleep(3)
        return {"id": vc_id, "name": name, "status": "FAILED", "stage": self._classify_failure_stage(last_error), "network": network, "error": last_error, "returncode": None if proc is None else proc.returncode}

    def collect_all(self) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        enabled = self.enabled_vcenters()
        if not enabled:
            raise PowerCLICollectionError("활성화된 vCenter가 없습니다. 연동정보 관리에서 먼저 등록하세요.")
        results = [self.run_one(entry) for entry in enabled]
        records: list[dict[str, Any]] = []
        success_scopes: list[str] = []
        failed_scopes: dict[str, str] = {}
        files: list[dict[str, Any]] = []
        for result in results:
            vc_id = str(result.get("id") or "UNKNOWN")
            if result.get("status") == "SUCCESS":
                success_scopes.append(vc_id)
                records.extend(result.get("records") or [])
                files.append({
                    "scope": vc_id,
                    "rows": int(result.get("row_count") or 0),
                    "json_file": result.get("json_file"),
                    "xlsx_file": result.get("xlsx_file"),
                })
            else:
                failed_scopes[vc_id] = str(result.get("error") or "PowerCLI collection failed")
        if not records:
            raise PowerCLICollectionError("정상 수집된 vCenter VM 데이터가 없습니다.")
        return records, {
            "mode": "POWERCLI",
            "success_scopes": success_scopes,
            "failed_scopes": failed_scopes,
            "results": [{k: v for k, v in item.items() if k != "records"} for item in results],
            "files": files,
            "failed_files": [],
        }

    def test_one(self, entry: dict[str, Any]) -> dict[str, Any]:
        result = self.run_one(entry, test_run=True)
        result.pop("records", None)
        result["json_created"] = bool(result.get("json_file"))
        result["xlsx_created"] = bool(result.get("xlsx_file"))
        for key in ("json_file", "xlsx_file"):
            value = result.get(key)
            if value:
                try:
                    Path(str(value)).unlink(missing_ok=True)
                except OSError:
                    LOGGER.warning("PowerCLI test artifact cleanup failed: %s", value)
            result[key] = None
        return result

    def test_all(self) -> dict[str, Any]:
        enabled = self.enabled_vcenters()
        if not enabled:
            raise PowerCLICollectionError("테스트할 활성 vCenter가 없습니다.")
        results = [self.test_one(entry) for entry in enabled]
        success = sum(1 for item in results if item.get("status") == "SUCCESS")
        return {
            "status": "SUCCESS" if success == len(results) else "PARTIAL_SUCCESS" if success else "FAILED",
            "total": len(results),
            "success": success,
            "failed": len(results) - success,
            "results": results,
        }


class VCenterSnapshotFileCollector:
    """Fallback loader for PowerCLI-exported JSON/XLSX snapshots."""

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def collect(self, files: list[Path] | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        incoming = self.config.resolve(self.config.rvtools.get("incoming_dir", "data/incoming/vcenter"))
        files = files or sorted([*incoming.glob("*.json"), *incoming.glob("*.xlsx")])
        if not files:
            raise PowerCLICollectionError(f"vCenter 스냅샷 파일이 없습니다: {incoming}")
        records: list[dict[str, Any]] = []
        metadata: list[dict[str, Any]] = []
        failed: dict[str, str] = {}
        for path in files:
            try:
                if path.suffix.lower() == ".json":
                    data = json.loads(path.read_text(encoding="utf-8-sig"))
                    rows = [data] if isinstance(data, dict) else data
                    rows = [dict(item) for item in rows if isinstance(item, dict)]
                else:
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb["vInfo"] if "vInfo" in wb.sheetnames else wb.active
                    headers = [str(value).strip() if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                    rows = []
                    for values in ws.iter_rows(min_row=2, values_only=True):
                        if not any(value is not None and str(value).strip() for value in values):
                            continue
                        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header})
                    wb.close()
                if not rows:
                    raise PowerCLICollectionError("데이터 행이 없습니다.")
                missing = sorted(_REQUIRED - set(rows[0]))
                if missing:
                    raise PowerCLICollectionError(f"필수 필드 누락: {', '.join(missing)}")
                scope = str(rows[0].get("VI SDK Server") or path.stem)
                for row in rows:
                    row.setdefault("_source_file", str(path))
                    row.setdefault("_vcenter_scope", str(row.get("VI SDK Server") or scope))
                records.extend(rows)
                metadata.append({"file": str(path), "rows": len(rows), "scope": scope})
            except Exception as exc:
                failed[path.stem] = str(exc)
        if not records:
            raise PowerCLICollectionError("정상 처리된 vCenter 스냅샷 파일이 없습니다.")
        return records, {"mode": "FILE_ONLY", "success_scopes": [item["scope"] for item in metadata], "failed_scopes": failed, "files": metadata, "failed_files": []}
