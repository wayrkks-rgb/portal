from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import AppConfig


class ITSMFileCollectionError(RuntimeError):
    pass


class ITSMFileCollector:
    """Load ITSM fallback data from the newest XLSX or CSV file."""

    def __init__(self, config: AppConfig) -> None:
        self.config = config
        self.last_metadata: dict[str, Any] = {}

    def collect(self, files: list[Path] | None = None) -> list[dict[str, Any]]:
        incoming = self.config.resolve(self.config.itsm.get("incoming_dir", "data/incoming/itsm"))
        if files is None:
            candidates = [*incoming.glob("*.xlsx"), *incoming.glob("*.csv")]
            files = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[:1]
        if not files:
            raise ITSMFileCollectionError(f"ITSM 파일이 없습니다: {incoming}")
        path = files[0]
        if path.suffix.lower() == ".csv":
            records = self._read_csv(path)
        elif path.suffix.lower() == ".xlsx":
            records = self._read_xlsx(path)
        else:
            raise ITSMFileCollectionError(f"지원하지 않는 파일 형식입니다: {path.suffix}")
        if not records:
            raise ITSMFileCollectionError("ITSM 파일 데이터가 0건입니다.")
        self.last_metadata = {"mode": "FILE_ONLY", "file": str(path), "rows": len(records)}
        return records

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            return [{str(k).strip().upper(): v for k, v in row.items()} for row in csv.DictReader(stream)]

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = str(self.config.itsm.get("sheet_name", "Sheet1"))
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        header_row = int(self.config.itsm.get("header_row", 1))
        header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
        headers = [str(value).strip().upper() if value is not None else "" for value in header_values]
        records: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            records.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header})
        workbook.close()
        return records
