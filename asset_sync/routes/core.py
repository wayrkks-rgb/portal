from __future__ import annotations

import csv
import io
import json
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from flask import Blueprint, jsonify, render_template, request, send_file, session

from ..config import AppConfig
from ..db.sqlite_manager import SQLiteManager
from ..repositories import AssetRepository
from ..services import (
    AutomatedReportService, ChangeSyncService, DailyComparisonService, DashboardService, ExportService,
    IntegratedDashboardService, PeriodService, ReconciliationExceptionService,
    ReconciliationService, VMResourceUsageExportService,
)
from ..web_common import admin_required, login_required


def create_core_blueprint(cfg: AppConfig, manager: SQLiteManager) -> Blueprint:
    bp = Blueprint("asset_sync_core", __name__)

    @bp.route("/asset-sync")
    @login_required
    def page() -> Any:
        return render_template("main.html", user=session["user"], page="asset_sync")

    @bp.route("/api/health")
    @bp.route("/api/asset-sync/health")
    @login_required
    def health() -> Any:
        try:
            with manager.connect() as conn:
                conn.execute("SELECT 1").fetchone()
            return jsonify({"status": "UP", "database": str(cfg.database_path)})
        except Exception as exc:
            return jsonify({"status": "DOWN", "error": str(exc)}), 500

    @bp.route("/api/dashboard/summary")
    @bp.route("/api/asset-sync/dashboard")
    @login_required
    def dashboard_summary() -> Any:
        try:
            with manager.connect() as conn:
                result = IntegratedDashboardService(AssetRepository(conn)).summary(
                    start=request.args.get("start"),
                    end=request.args.get("end"),
                    detail_limit=min(int(request.args.get("limit", 500)), 5000),
                )
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/dashboard/legacy")
    @login_required
    def legacy_dashboard_summary() -> Any:
        with manager.connect() as conn:
            return jsonify(DashboardService(AssetRepository(conn)).summary())

    @bp.route("/api/collection-runs")
    @bp.route("/api/asset-sync/collection-runs")
    @login_required
    def collection_runs() -> Any:
        limit = min(int(request.args.get("limit", 100)), 1000)
        with manager.connect() as conn:
            return jsonify(AssetRepository(conn).collection_runs(limit))

    @bp.route("/api/asset-sync/daily-comparison")
    @login_required
    def daily_comparison() -> Any:
        source = request.args.get("source", "ITSM").upper()
        limit = min(int(request.args.get("limit", 2000)), 10000)
        try:
            with manager.connect() as conn:
                return jsonify(DailyComparisonService(cfg, AssetRepository(conn)).latest(source, limit))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/changes")
    @bp.route("/api/asset-sync/changes")
    @login_required
    def changes() -> Any:
        source = request.args.get("source")
        limit = min(int(request.args.get("limit", 500)), 10000)
        with manager.connect() as conn:
            return jsonify(
                AssetRepository(conn).changes(
                    source=source,
                    limit=limit,
                    start=request.args.get("start"),
                    end=request.args.get("end"),
                )
            )

    @bp.route("/api/reconciliation")
    @bp.route("/api/asset-sync/reconciliation")
    @login_required
    def reconciliation() -> Any:
        limit = min(int(request.args.get("limit", 1000)), 10000)
        status = request.args.get("status")
        with manager.connect() as conn:
            rows = AssetRepository(conn).reconciliation(limit=limit, status=status)
            for row in rows:
                row["drifts"] = json.loads(row.pop("drift_json") or "[]")
            return jsonify(rows)

    @bp.route("/api/reconciliation/<int:result_id>")
    @login_required
    def reconciliation_detail(result_id: int) -> Any:
        with manager.connect() as conn:
            row = conn.execute("SELECT * FROM reconciliation_result WHERE id=?", (result_id,)).fetchone()
            if not row:
                return jsonify({"error": "결과가 없습니다."}), 404
            data = dict(row)
            data["drifts"] = json.loads(data.pop("drift_json") or "[]")
            return jsonify(data)

    @bp.route("/api/asset-sync/reconcile", methods=["POST"])
    @admin_required
    def run_reconciliation() -> Any:
        with manager.connect() as conn:
            return jsonify(ReconciliationService(cfg, AssetRepository(conn)).reconcile_latest())

    @bp.route("/api/asset-sync/sync-results")
    @login_required
    def sync_results() -> Any:
        limit = min(int(request.args.get("limit", 1000)), 10000)
        with manager.connect() as conn:
            return jsonify(ChangeSyncService(cfg, AssetRepository(conn)).latest_results(limit))

    @bp.route("/api/asset-sync/evaluate-sync", methods=["POST"])
    @admin_required
    def evaluate_sync() -> Any:
        payload = request.get_json() or {}
        if not payload.get("start") or not payload.get("end"):
            return jsonify({"error": "start/end가 필요합니다."}), 400
        with manager.connect() as conn:
            return jsonify(ChangeSyncService(cfg, AssetRepository(conn)).evaluate(payload["start"], payload["end"]))

    @bp.route("/api/asset-sync/period-summary")
    @login_required
    def period_summary() -> Any:
        source = request.args.get("source", "ITSM").upper()
        mode = request.args.get("mode", "custom")
        with manager.connect() as conn:
            service = PeriodService(AssetRepository(conn))
            if mode == "daily":
                return jsonify(service.daily(source, request.args["date"]))
            if mode == "weekly":
                return jsonify(service.weekly(source, request.args["end_date"]))
            if mode == "monthly":
                return jsonify(service.monthly(source, int(request.args["year"]), int(request.args["month"])))
            return jsonify(service.summary(source, request.args["start"], request.args["end"]))

    @bp.route("/api/asset-sync/reconciliation/candidates")
    @login_required
    def reconciliation_candidates() -> Any:
        with manager.connect() as conn:
            rows = IntegratedDashboardService(AssetRepository(conn)).exception_candidates(
                min(int(request.args.get("limit", 5000)), 10000)
            )
        return jsonify(rows)

    @bp.route("/api/asset-sync/exceptions")
    @login_required
    def reconciliation_exceptions() -> Any:
        include_inactive = request.args.get("include_inactive", "1") != "0"
        with manager.connect() as conn:
            return jsonify(ReconciliationExceptionService(AssetRepository(conn)).list(include_inactive))

    @bp.route("/api/asset-sync/exceptions", methods=["POST"])
    @admin_required
    def create_reconciliation_exceptions() -> Any:
        payload = request.get_json(silent=True) or {}
        items = payload.get("items") if isinstance(payload.get("items"), list) else [payload]
        user_id = str(session.get("user", {}).get("username") or session.get("user", {}).get("name") or "ADMIN")
        try:
            with manager.connect() as conn:
                result = ReconciliationExceptionService(AssetRepository(conn)).create_many(items, user_id)
            code = 201 if result["created_count"] else 400
            return jsonify(result), code
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/exceptions/<int:exception_id>", methods=["DELETE"])
    @admin_required
    def deactivate_reconciliation_exception(exception_id: int) -> Any:
        user_id = str(session.get("user", {}).get("username") or session.get("user", {}).get("name") or "ADMIN")
        with manager.connect() as conn:
            changed = ReconciliationExceptionService(AssetRepository(conn)).deactivate(exception_id, user_id)
        if not changed:
            return jsonify({"error": "활성 예외처리를 찾을 수 없습니다."}), 404
        return jsonify({"status": "SUCCESS", "exception_id": exception_id})

    @bp.route("/api/asset-sync/exceptions/import", methods=["POST"])
    @admin_required
    def import_reconciliation_exceptions() -> Any:
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"error": "CSV 또는 XLSX 파일이 필요합니다."}), 400
        try:
            items = _read_exception_upload(upload.filename, upload.read())
            user_id = str(session.get("user", {}).get("username") or session.get("user", {}).get("name") or "ADMIN")
            with manager.connect() as conn:
                result = ReconciliationExceptionService(AssetRepository(conn)).create_many(items, user_id)
            return jsonify(result), 201 if result["created_count"] else 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/resource-usage")
    @login_required
    def resource_usage_summary() -> Any:
        start = request.args.get("start") or (date.today() - timedelta(days=30)).isoformat()
        end = request.args.get("end") or (date.today() - timedelta(days=1)).isoformat()
        try:
            with manager.connect() as conn:
                result = VMResourceUsageExportService(cfg, AssetRepository(conn)).summary(
                    start, end,
                    vcenter_id=request.args.get("vcenter_id") or None,
                    cluster_name=request.args.get("cluster_name") or None,
                    esxi_host=request.args.get("esxi_host") or None,
                )
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/resource-usage/export")
    @login_required
    def resource_usage_export() -> Any:
        start = request.args.get("start") or (date.today() - timedelta(days=30)).isoformat()
        end = request.args.get("end") or (date.today() - timedelta(days=1)).isoformat()
        try:
            with manager.connect() as conn:
                path = VMResourceUsageExportService(cfg, AssetRepository(conn)).export_xlsx(
                    start, end, cfg.resolve("data/export/resource_usage"),
                    vcenter_id=request.args.get("vcenter_id") or None,
                    cluster_name=request.args.get("cluster_name") or None,
                    esxi_host=request.args.get("esxi_host") or None,
                )
            return send_file(path, as_attachment=True, download_name=path.name)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/resource-usage/import", methods=["POST"])
    @admin_required
    def import_resource_usage() -> Any:
        upload = request.files.get("file")
        if not upload or not upload.filename:
            return jsonify({"error": "VM_ResourceUsageExport 결과 파일이 필요합니다."}), 400
        target = cfg.resolve("data/temp/resource_usage_import") / Path(upload.filename).name
        target.parent.mkdir(parents=True, exist_ok=True)
        upload.save(target)
        try:
            with manager.connect() as conn:
                result = VMResourceUsageExportService(cfg, AssetRepository(conn)).import_file(
                    target, request.form.get("stat_date") or None
                )
            return jsonify(result)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400
        finally:
            target.unlink(missing_ok=True)

    @bp.route("/api/asset-sync/reports/<report_type>")
    @login_required
    def automated_report(report_type: str) -> Any:
        try:
            with manager.connect() as conn:
                path = AutomatedReportService(
                    AssetRepository(conn), cfg.resolve("data/export/automated_reports")
                ).generate(report_type, request.args.get("start"), request.args.get("end"))
            return send_file(path, as_attachment=True, download_name=path.name)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @bp.route("/api/asset-sync/export")
    @login_required
    def export_results() -> Any:
        with manager.connect() as conn:
            path = ExportService(AssetRepository(conn), cfg.resolve("data/export")).export_current()
        return send_file(path, as_attachment=True, download_name=path.name)

    return bp


def _read_exception_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        rows = [dict(zip(headers, row)) for row in values]
    else:
        raise ValueError("지원 파일은 CSV 또는 XLSX입니다.")
    aliases = {
        "예외유형": "exception_type", "EXCEPTION_TYPE": "exception_type",
        "CM_ID": "cm_id", "ITSM_ID": "cm_id",
        "VCENTER_KEY": "rv_asset_key", "RV_ASSET_KEY": "rv_asset_key", "VM_UUID": "rv_asset_key",
        "서버명": "server_name", "SERVER_NAME": "server_name",
        "사유": "reason", "REASON": "reason",
        "시작일": "valid_from", "VALID_FROM": "valid_from",
        "종료일": "valid_to", "VALID_TO": "valid_to",
    }
    result: list[dict[str, Any]] = []
    for source in rows:
        item: dict[str, Any] = {}
        for key, value in source.items():
            alias = aliases.get(str(key or "").strip().upper()) or aliases.get(str(key or "").strip())
            if alias:
                item[alias] = value
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    if not result:
        raise ValueError("등록할 예외 데이터가 없습니다.")
    return result
