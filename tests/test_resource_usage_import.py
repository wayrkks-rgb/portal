from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from asset_sync.config import AppConfig
from asset_sync.db.sqlite_manager import SQLiteManager
from asset_sync.repositories import AssetRepository
from asset_sync.services.resource_usage_service import VMResourceUsageExportService


def test_vm_resource_usage_export_json_import(tmp_path: Path) -> None:
    cfg = AppConfig(root_dir=tmp_path, sqlite_path=Path("data/test.db"), rvtools={"resource_usage": {"enabled": False}})
    manager = SQLiteManager(cfg.database_path)
    manager.initialize()
    source = tmp_path / "VM_ResourceUsageExport.json"
    source.write_text(json.dumps([
        {"ENTITY_TYPE": "ESXI", "VCENTER_ID": "VC-A", "ESXI_HOST": "HOST-A", "CPU_MAX": 81.2, "CPU_AVG": 42.1, "MEM_MAX": 72.4, "MEM_AVG": 55.0},
        {"ENTITY_TYPE": "VM", "VCENTER_ID": "VC-A", "ESXI_HOST": "HOST-A", "VM_UUID": "UUID-1", "VM_NAME": "VM-1", "CPU_MAX": 30, "CPU_AVG": 10, "MEM_MAX": 65, "MEM_AVG": 50},
    ]), encoding="utf-8")
    with manager.connect() as conn:
        result = VMResourceUsageExportService(cfg, AssetRepository(conn)).import_file(source, "2026-07-23")
        assert result["count"] == 2
        count = conn.execute("SELECT COUNT(*) AS cnt FROM vcenter_resource_daily WHERE stat_date='2026-07-23'").fetchone()["cnt"]
        assert count == 2


def test_resource_usage_summary_and_export_from_daily_batch(tmp_path: Path) -> None:
    from asset_sync.services.collection_service import CollectionService

    cfg = AppConfig(
        root_dir=tmp_path,
        sqlite_path=Path("data/test_batch.db"),
        itsm={
            "collection_mode": "DEMO", "memory_unit": "GB", "cpu_compare_field": "CM_CPU_CORE_CNT",
            "memory_field": "CM_MEMORY", "os_eos_field": "OS_EOS_DATE", "tracked_fields": [], "ignore_fields": [],
            "owner_field": "CM_WOR_MNG_EMP_ID",
        },
        rvtools={
            "collection_mode": "DEMO", "hostname_suffixes": [".example.invalid"],
            "power_on_value": "poweredon", "exclude_templates": True, "exclude_srm_placeholders": True,
            "resource_usage": {"enabled": True},
        },
        matching={"use_identity_map_first": True, "auto_remember_exact_match": True, "memory_tolerance_mb": 1},
        quality={"minimum_itsm_records": 1, "minimum_rvtools_records": 1, "eos_near_days": 180},
    )
    manager = SQLiteManager(cfg.database_path)
    manager.initialize()
    batch = CollectionService(cfg, manager).run_daily(demo=True)
    period = batch["resource_usage"]["period_end"]
    with manager.connect() as conn:
        service = VMResourceUsageExportService(cfg, AssetRepository(conn))
        result = service.summary(period, period)
        assert result["summary"]["host_count"] >= 1
        assert result["summary"]["vm_count"] >= 1
        assert result["vms"][0]["allocated_cpu_cores"] is not None
        assert result["vms"][0]["allocated_memory_mb"] is not None
        target = service.export_xlsx(period, period, tmp_path / "export")
        assert target.exists()
        workbook = load_workbook(target, read_only=True)
        assert workbook.sheetnames == ["HostResourceUsage", "VMsResource", "VMChangeHistory"]
        workbook.close()
